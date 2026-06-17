"""Fetch a web page, extract readable text, and save it as a Markdown source doc.

Usage:
    python scrape_url.py "https://example.com/article"
    python scrape_url.py "https://example.com/article" --title "Custom Title"
    python scrape_url.py "https://example.com/article" --output documents/custom_name.md

For CSV, TSV, public Google Sheets, and HTML tables, rows are saved as
Markdown sections so column/value relationships survive chunking.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from io import BytesIO, StringIO
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qs, urlparse

import requests
from bs4 import BeautifulSoup, Tag


DOCUMENTS_DIR = Path("documents")
DEFAULT_TIMEOUT_SECONDS = 20
USER_AGENT = (
    "Mozilla/5.0 (compatible; UniversityGuideBot/1.0; "
    "+https://example.com/student-project)"
)


def fetch_url(url: str) -> requests.Response:
    response = requests.get(
        url,
        headers={"User-Agent": USER_AGENT},
        timeout=DEFAULT_TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    return response


def google_sheets_csv_url(url: str) -> str | None:
    parsed = urlparse(url)
    if parsed.netloc.lower() != "docs.google.com":
        return None

    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return None

    query = parse_qs(parsed.query)
    fragment = parse_qs(parsed.fragment)
    gid = query.get("gid", fragment.get("gid", ["0"]))[0]
    sheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def looks_like_tabular_response(url: str, content_type: str) -> bool:
    parsed = urlparse(url)
    path = parsed.path.lower()
    content_type = content_type.lower()

    return (
        path.endswith((".csv", ".tsv", ".xlsx"))
        or "text/csv" in content_type
        or "application/csv" in content_type
        or "tab-separated-values" in content_type
        or "spreadsheetml" in content_type
    )


def title_from_url(url: str) -> str:
    parsed = urlparse(url)
    path_name = Path(parsed.path).stem
    if path_name:
        return clean_inline_text(path_name.replace("-", " ").replace("_", " ")).title()

    return parsed.netloc or "Tabular Source"


def extract_title(soup: BeautifulSoup, fallback_url: str) -> str:
    if soup.title and soup.title.string:
        title = clean_inline_text(soup.title.string)
        if title:
            return title

    heading = soup.find("h1")
    if heading:
        title = clean_inline_text(heading.get_text(" ", strip=True))
        if title:
            return title

    parsed = urlparse(fallback_url)
    return parsed.netloc or "Untitled Source"


def clean_soup(soup: BeautifulSoup) -> None:
    for tag in soup(
        [
            "script",
            "style",
            "noscript",
            "svg",
            "canvas",
            "iframe",
            "form",
            "button",
            "input",
            "select",
            "textarea",
            "nav",
            "footer",
            "header",
            "aside",
        ]
    ):
        tag.decompose()

    for selector in [
        "[aria-hidden='true']",
        ".cookie",
        ".cookies",
        ".advertisement",
        ".ad",
        ".ads",
        ".social",
        ".share",
        ".newsletter",
        ".subscribe",
    ]:
        for tag in soup.select(selector):
            tag.decompose()


def choose_content_root(soup: BeautifulSoup) -> Tag:
    for selector in ["main", "article", "[role='main']", ".content", "#content"]:
        match = soup.select_one(selector)
        if match and len(clean_inline_text(match.get_text(" ", strip=True))) > 200:
            return match

    body = soup.body
    if body:
        return body

    return soup


def clean_inline_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def decode_text_response(response: requests.Response) -> str:
    if response.encoding:
        return response.text

    return response.content.decode("utf-8-sig", errors="replace")


def parse_delimited_rows(text: str, delimiter: str | None = None) -> tuple[list[str], list[list[str]]]:
    sample = text[:4096]
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "\t" if "\t" in sample else ","

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    raw_rows = [
        [clean_inline_text(cell) for cell in row]
        for row in reader
        if any(clean_inline_text(cell) for cell in row)
    ]

    if not raw_rows:
        return [], []

    headers = normalize_headers(raw_rows[0])
    rows = [normalize_row(row, len(headers)) for row in raw_rows[1:]]
    return headers, rows


def parse_xlsx_rows(content: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "This looks like an .xlsx spreadsheet. Install openpyxl or export it as CSV."
        ) from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows: list[list[str]] = []

    for row in sheet.iter_rows(values_only=True):
        values = [clean_inline_text("" if cell is None else str(cell)) for cell in row]
        if any(values):
            raw_rows.append(values)

    if not raw_rows:
        return [], []

    headers = normalize_headers(raw_rows[0])
    rows = [normalize_row(row, len(headers)) for row in raw_rows[1:]]
    return headers, rows


def normalize_headers(raw_headers: list[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index, header in enumerate(raw_headers, start=1):
        header = clean_inline_text(header) or f"Column {index}"
        count = seen.get(header.casefold(), 0) + 1
        seen[header.casefold()] = count
        if count > 1:
            header = f"{header} {count}"
        headers.append(header)

    return headers


def normalize_row(row: list[str], width: int) -> list[str]:
    row = [clean_inline_text(cell) for cell in row]
    if len(row) < width:
        row = row + [""] * (width - len(row))
    return row[:width]


def row_heading(headers: list[str], row: list[str], row_number: int) -> str:
    preferred_terms = [
        "hackathon",
        "event",
        "project",
        "name",
        "title",
        "source",
        "url",
    ]

    for term in preferred_terms:
        for header, value in zip(headers, row):
            if value and term in header.casefold():
                return f"Row {row_number}: {value}"

    for value in row:
        if value:
            return f"Row {row_number}: {value}"

    return f"Row {row_number}"


def rows_to_markdown(headers: list[str], rows: list[list[str]], table_title: str | None = None) -> str:
    if not headers or not rows:
        return ""

    blocks: list[str] = []
    if table_title:
        blocks.append(f"## {table_title}")

    for index, row in enumerate(rows, start=1):
        row = normalize_row(row, len(headers))
        fields = [
            f"- {header}: {value}"
            for header, value in zip(headers, row)
            if clean_inline_text(value)
        ]
        if not fields:
            continue

        blocks.append(f"### {row_heading(headers, row, index)}")
        blocks.extend(fields)

    return "\n\n".join(blocks)


def html_table_to_rows(table: Tag) -> tuple[list[str], list[list[str]]]:
    parsed_rows: list[list[str]] = []

    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        values = [clean_inline_text(cell.get_text(" ", strip=True)) for cell in cells]
        if any(values):
            parsed_rows.append(values)

    if len(parsed_rows) < 2:
        return [], []

    headers = normalize_headers(parsed_rows[0])
    rows = [normalize_row(row, len(headers)) for row in parsed_rows[1:]]
    return headers, rows


def markdown_from_tables(soup: BeautifulSoup) -> str:
    table_blocks: list[str] = []

    for index, table in enumerate(soup.find_all("table"), start=1):
        headers, rows = html_table_to_rows(table)
        if not headers or not rows:
            continue

        caption = table.find("caption")
        title = clean_inline_text(caption.get_text(" ", strip=True)) if caption else f"Table {index}"
        table_markdown = rows_to_markdown(headers, rows, title)
        if table_markdown:
            table_blocks.append(table_markdown)

    return "\n\n".join(table_blocks)


def markdown_from_element(element: Tag) -> str:
    blocks: list[str] = []

    for node in element.descendants:
        if not isinstance(node, Tag):
            continue

        name = node.name.lower()
        text = clean_inline_text(node.get_text(" ", strip=True))
        if not text:
            continue

        if name in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            level = min(int(name[1]), 6)
            blocks.append(f"{'#' * level} {text}")
        elif name == "li":
            blocks.append(f"- {text}")
        elif name == "blockquote":
            blocks.append(f"> {text}")
        elif name in {"p", "pre"}:
            blocks.append(text)

    if not blocks:
        text = element.get_text("\n", strip=True)
        blocks = [clean_inline_text(line) for line in text.splitlines()]

    return "\n\n".join(deduplicate_blocks(blocks))


def deduplicate_blocks(blocks: Iterable[str]) -> list[str]:
    cleaned: list[str] = []
    seen_recent: set[str] = set()

    for block in blocks:
        block = clean_inline_text(block)
        if len(block) < 2:
            continue

        normalized = block.casefold()
        if normalized in seen_recent:
            continue

        cleaned.append(block)
        seen_recent.add(normalized)

        if len(seen_recent) > 250:
            seen_recent = {item.casefold() for item in cleaned[-100:]}

    return cleaned


def slugify(value: str, max_length: int = 80) -> str:
    value = value.lower()
    value = re.sub(r"https?://", "", value)
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = value.strip("-")
    return (value[:max_length].strip("-") or "source")


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    parent = path.parent

    counter = 2
    while True:
        candidate = parent / f"{stem}-{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def build_markdown(
    url: str,
    title: str,
    body_text: str,
    capture_method: str,
    source_type: str = "web_page",
) -> str:
    collected_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    safe_title = title.replace('"', "'")
    safe_url = url.replace('"', "%22")
    safe_source_type = source_type.replace('"', "'")

    return (
        "---\n"
        f'title: "{safe_title}"\n'
        f'source_url: "{safe_url}"\n'
        f'source_type: "{safe_source_type}"\n'
        f'collected_at: "{collected_at}"\n'
        f'capture_method: "{capture_method}"\n'
        "---\n\n"
        f"# {title}\n\n"
        f"{body_text.strip()}\n"
    )


def scrape_url_to_markdown(url: str, output: Path | None, title_override: str | None) -> Path:
    google_csv_url = google_sheets_csv_url(url)
    fetch_target = google_csv_url or url
    response = fetch_url(fetch_target)
    content_type = response.headers.get("content-type", "")
    source_type = "web_page"
    capture_method = "scrape_url.py"

    if google_csv_url or looks_like_tabular_response(fetch_target, content_type):
        source_type = "tabular_data"
        capture_method = "scrape_url.py:tabular"
        title = title_override or title_from_url(url)

        if urlparse(fetch_target).path.lower().endswith(".xlsx") or "spreadsheetml" in content_type.lower():
            headers, rows = parse_xlsx_rows(response.content)
        else:
            text = decode_text_response(response)
            delimiter = "\t" if "tab-separated-values" in content_type.lower() else None
            headers, rows = parse_delimited_rows(text, delimiter=delimiter)

        body_text = rows_to_markdown(headers, rows)
    else:
        if "html" not in content_type.lower():
            raise ValueError(
                f"Expected an HTML or tabular source, but got content type: "
                f"{content_type or 'unknown'}"
            )

        soup = BeautifulSoup(response.text, "html.parser")
        clean_soup(soup)

        title = title_override or extract_title(soup, url)
        table_text = markdown_from_tables(soup)
        if table_text:
            source_type = "html_table"
            capture_method = "scrape_url.py:html_table"
            body_text = table_text
        else:
            content_root = choose_content_root(soup)
            body_text = markdown_from_element(content_root)

    if len(body_text) < 200:
        raise ValueError(
            "Extracted less than 200 characters. This page may be JavaScript-rendered, "
            "blocked, empty, or better collected manually."
        )

    DOCUMENTS_DIR.mkdir(exist_ok=True)

    if output is None:
        parsed = urlparse(url)
        source_hint = f"{parsed.netloc}-{title}"
        output = DOCUMENTS_DIR / f"{slugify(source_hint)}.md"
    elif output.suffix.lower() != ".md":
        output = output.with_suffix(".md")

    output = unique_path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        build_markdown(url, title, body_text, capture_method, source_type),
        encoding="utf-8",
    )

    return output


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape one URL and save cleaned text as a Markdown file in documents/."
    )
    parser.add_argument("url", help="The web URL to scrape.")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        help="Optional output path. Defaults to documents/<page-title>.md.",
    )
    parser.add_argument(
        "--title",
        help="Optional title override for the saved Markdown metadata and heading.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)

    try:
        saved_path = scrape_url_to_markdown(args.url, args.output, args.title)
    except Exception as exc:
        print(f"Failed to scrape URL: {exc}", file=sys.stderr)
        return 1

    print(f"Saved scraped document to: {saved_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

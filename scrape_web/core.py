"""Fetch a web page, extract readable text, and save it as a Markdown source doc.

Usage:
    python -m scrape_web "https://example.com/article"
    python -m scrape_web "https://example.com/article" --title "Custom Title"
    python -m scrape_web "https://example.com/article" --output documents/custom_name.md
    python -m scrape_web "https://www.reddit.com/r/hackathons/comments/..." --reddit-max-comments 40
    python -m scrape_web "https://us.allhackathons.com/themes/university/" --page-end 18 --output documents/allhackathons_university.md
    python -m scrape_web "https://us.allhackathons.com/themes/university/" --page-end 18 --follow-read-more --output documents/allhackathons_university_details.md

For CSV, TSV, public Google Sheets, and HTML tables, rows are saved as
Markdown sections so column/value relationships survive chunking.
For Reddit thread URLs, the script uses Reddit's public JSON representation
and saves the original post plus readable comments with permalinks.
For paginated pages, the script saves all requested pages into one Markdown file.
Use --follow-read-more for listing/card pages to save linked detail pages instead
of teaser text from the listing cards.
Use --include-visible-text to preserve short visible UI text such as tags,
badges, statuses, and action labels.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from dataclasses import dataclass
from html import unescape
from io import BytesIO, StringIO
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qs, parse_qsl, urlencode, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup, Tag
from dotenv import load_dotenv


DOCUMENTS_DIR = Path("documents")
DEFAULT_TIMEOUT_SECONDS = 20
USER_AGENT = (
    "Mozilla/5.0 (compatible; UniversityGuideBot/1.0; "
    "+https://example.com/student-project)"
)
REDDIT_USER_AGENT = "python:university-guide-rag:v1.0 (student research project)"


@dataclass
class ScrapedContent:
    title: str
    body_text: str
    source_type: str
    capture_method: str
    extra_metadata: dict[str, str]


def fetch_url(url: str, headers: dict[str, str] | None = None) -> requests.Response:
    request_headers = {"User-Agent": USER_AGENT}
    if headers:
        request_headers.update(headers)

    response = requests.get(
        url,
        headers=request_headers,
        timeout=DEFAULT_TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    return response


def google_sheets_csv_url(url: str) -> str | None:
    parsed = urlparse(url)
    if parsed.netloc.lower() != "docs.google.com":
        return None

    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return None

    query = parse_qs(parsed.query)
    fragment = parse_qs(parsed.fragment)
    gid = query.get("gid", fragment.get("gid", ["0"]))[0]
    sheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def looks_like_tabular_response(url: str, content_type: str) -> bool:
    parsed = urlparse(url)
    path = parsed.path.lower()
    content_type = content_type.lower()

    return (
        path.endswith((".csv", ".tsv", ".xlsx"))
        or "text/csv" in content_type
        or "application/csv" in content_type
        or "tab-separated-values" in content_type
        or "spreadsheetml" in content_type
    )


def is_reddit_url(url: str) -> bool:
    host = urlparse(url).netloc.lower()
    return host == "redd.it" or host == "reddit.com" or host.endswith(".reddit.com")


def reddit_json_url(url: str, max_comments: int, sort: str) -> str:
    parsed = urlparse(url)
    host = parsed.netloc.lower()

    if host == "redd.it":
        post_id = parsed.path.strip("/").split("/")[0]
        if not post_id:
            raise ValueError("Could not find a Reddit post id in this redd.it URL.")
        path = f"/comments/{post_id}"
    else:
        path = parsed.path

    path = re.sub(r"/+$", "", path)
    path = re.sub(r"\.json$", "", path)

    if "/comments/" not in path:
        raise ValueError(
            "Reddit scraping expects a thread URL containing /comments/ or a redd.it link."
        )

    query = urlencode(
        {
            "limit": max_comments,
            "sort": sort,
            "raw_json": 1,
        }
    )
    return f"https://www.reddit.com{path}.json?{query}"


def reddit_thread_path(url: str) -> str:
    parsed = urlparse(url)
    host = parsed.netloc.lower()

    if host == "redd.it":
        post_id = parsed.path.strip("/").split("/")[0]
        if not post_id:
            raise ValueError("Could not find a Reddit post id in this redd.it URL.")
        return f"/comments/{post_id}"

    path = re.sub(r"/+$", "", parsed.path)
    path = re.sub(r"\.json$", "", path)
    if "/comments/" not in path:
        raise ValueError(
            "Reddit scraping expects a thread URL containing /comments/ or a redd.it link."
        )

    return path


def reddit_old_html_url(url: str, sort: str) -> str:
    path = reddit_thread_path(url)
    query = urlencode({"sort": sort})
    return f"https://old.reddit.com{path}/?{query}"


def reddit_oauth_json_url(url: str, max_comments: int, sort: str) -> str:
    path = reddit_thread_path(url)
    query = urlencode(
        {
            "limit": max_comments,
            "sort": sort,
            "raw_json": 1,
        }
    )
    return f"https://oauth.reddit.com{path}?{query}"


def reddit_oauth_token() -> str | None:
    client_id = os.environ.get("REDDIT_CLIENT_ID")
    client_secret = os.environ.get("REDDIT_CLIENT_SECRET")

    if not client_id or not client_secret:
        return None

    response = requests.post(
        "https://www.reddit.com/api/v1/access_token",
        auth=(client_id, client_secret),
        data={"grant_type": "client_credentials"},
        headers={"User-Agent": REDDIT_USER_AGENT},
        timeout=DEFAULT_TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    payload = response.json()
    token = payload.get("access_token")
    if not isinstance(token, str) or not token:
        raise ValueError("Reddit OAuth did not return an access token.")

    return token


def title_from_url(url: str) -> str:
    parsed = urlparse(url)
    path_name = Path(parsed.path).stem
    if path_name:
        return clean_inline_text(path_name.replace("-", " ").replace("_", " ")).title()

    return parsed.netloc or "Tabular Source"


def extract_title(soup: BeautifulSoup, fallback_url: str) -> str:
    if soup.title and soup.title.string:
        title = clean_inline_text(soup.title.string)
        if title:
            return title

    heading = soup.find("h1")
    if heading:
        title = clean_inline_text(heading.get_text(" ", strip=True))
        if title:
            return title

    parsed = urlparse(fallback_url)
    return parsed.netloc or "Untitled Source"


def clean_soup(soup: BeautifulSoup, keep_visible_chrome: bool = False) -> None:
    remove_tags = [
        "script",
        "style",
        "noscript",
        "svg",
        "canvas",
        "iframe",
        "form",
        "input",
        "select",
        "textarea",
    ]

    if not keep_visible_chrome:
        remove_tags.extend(["button", "nav", "footer", "header", "aside"])

    for tag in soup(remove_tags):
        tag.decompose()

    for selector in [
        "[aria-hidden='true']",
        ".cookie",
        ".cookies",
        ".advertisement",
        ".ad",
        ".ads",
        ".social",
        ".share",
        ".newsletter",
        ".subscribe",
    ]:
        for tag in soup.select(selector):
            tag.decompose()


def choose_content_root(soup: BeautifulSoup) -> Tag:
    for selector in ["main", "article", "[role='main']", ".content", "#content"]:
        match = soup.select_one(selector)
        if match and len(clean_inline_text(match.get_text(" ", strip=True))) > 200:
            return match

    body = soup.body
    if body:
        return body

    return soup


def clean_inline_text(text: str) -> str:
    text = unescape(text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def decode_text_response(response: requests.Response) -> str:
    if response.encoding:
        return response.text

    return response.content.decode("utf-8-sig", errors="replace")


def parse_delimited_rows(text: str, delimiter: str | None = None) -> tuple[list[str], list[list[str]]]:
    sample = text[:4096]
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "\t" if "\t" in sample else ","

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    raw_rows = [
        [clean_inline_text(cell) for cell in row]
        for row in reader
        if any(clean_inline_text(cell) for cell in row)
    ]

    if not raw_rows:
        return [], []

    headers = normalize_headers(raw_rows[0])
    rows = [normalize_row(row, len(headers)) for row in raw_rows[1:]]
    return headers, rows


def parse_xlsx_rows(content: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "This looks like an .xlsx spreadsheet. Install openpyxl or export it as CSV."
        ) from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    raw_rows: list[list[str]] = []

    for row in sheet.iter_rows(values_only=True):
        values = [clean_inline_text("" if cell is None else str(cell)) for cell in row]
        if any(values):
            raw_rows.append(values)

    if not raw_rows:
        return [], []

    headers = normalize_headers(raw_rows[0])
    rows = [normalize_row(row, len(headers)) for row in raw_rows[1:]]
    return headers, rows


def normalize_headers(raw_headers: list[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index, header in enumerate(raw_headers, start=1):
        header = clean_inline_text(header) or f"Column {index}"
        count = seen.get(header.casefold(), 0) + 1
        seen[header.casefold()] = count
        if count > 1:
            header = f"{header} {count}"
        headers.append(header)

    return headers


def normalize_row(row: list[str], width: int) -> list[str]:
    row = [clean_inline_text(cell) for cell in row]
    if len(row) < width:
        row = row + [""] * (width - len(row))
    return row[:width]


def row_heading(headers: list[str], row: list[str], row_number: int) -> str:
    preferred_terms = [
        "hackathon",
        "event",
        "project",
        "name",
        "title",
        "source",
        "url",
    ]

    for term in preferred_terms:
        for header, value in zip(headers, row):
            if value and term in header.casefold():
                return f"Row {row_number}: {value}"

    for value in row:
        if value:
            return f"Row {row_number}: {value}"

    return f"Row {row_number}"


def rows_to_markdown(headers: list[str], rows: list[list[str]], table_title: str | None = None) -> str:
    if not headers or not rows:
        return ""

    blocks: list[str] = []
    if table_title:
        blocks.append(f"## {table_title}")

    for index, row in enumerate(rows, start=1):
        row = normalize_row(row, len(headers))
        fields = [
            f"- {header}: {value}"
            for header, value in zip(headers, row)
            if clean_inline_text(value)
        ]
        if not fields:
            continue

        blocks.append(f"### {row_heading(headers, row, index)}")
        blocks.extend(fields)

    return "\n\n".join(blocks)


def reddit_permalink(permalink: str | None) -> str:
    if not permalink:
        return ""

    if permalink.startswith("http"):
        return permalink

    return f"https://www.reddit.com{permalink}"


def reddit_text(value: object) -> str:
    if not isinstance(value, str):
        return ""

    value = clean_inline_text(value)
    if value.casefold() in {"[deleted]", "[removed]"}:
        return ""

    return value


def reddit_body_text(value: object) -> str:
    if not isinstance(value, str):
        return ""

    value = value.replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ")
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in value.split("\n")]

    cleaned_lines: list[str] = []
    blank_count = 0
    for line in lines:
        if line:
            cleaned_lines.append(line)
            blank_count = 0
        else:
            blank_count += 1
            if blank_count <= 1:
                cleaned_lines.append("")

    cleaned = "\n".join(cleaned_lines).strip()
    if cleaned.casefold() in {"[deleted]", "[removed]"}:
        return ""

    return cleaned


def reddit_metadata_lines(data: dict, include_authors: bool) -> list[str]:
    lines: list[str] = []

    if include_authors:
        author = reddit_text(data.get("author"))
        if author:
            lines.append(f"- Author: u/{author}")

    score = data.get("score")
    if isinstance(score, int):
        lines.append(f"- Score: {score}")

    permalink = reddit_permalink(data.get("permalink"))
    if permalink:
        lines.append(f"- Permalink: {permalink}")

    return lines


def append_reddit_comment(
    child: dict,
    blocks: list[str],
    path: str,
    depth: int,
    remaining: list[int],
    include_authors: bool,
) -> None:
    if remaining[0] <= 0 or child.get("kind") != "t1":
        return

    data = child.get("data", {})
    body = reddit_body_text(data.get("body"))
    if not body:
        return

    remaining[0] -= 1
    heading_level = min(3 + depth, 6)
    section = [f"{'#' * heading_level} Comment {path}"]
    section.extend(reddit_metadata_lines(data, include_authors))
    section.append(body)
    blocks.append("\n\n".join(section))

    replies = data.get("replies")
    if remaining[0] <= 0 or not isinstance(replies, dict):
        return

    reply_children = replies.get("data", {}).get("children", [])
    reply_index = 1
    for reply in reply_children:
        if reply.get("kind") != "t1":
            continue

        append_reddit_comment(
            reply,
            blocks,
            f"{path}.{reply_index}",
            depth + 1,
            remaining,
            include_authors,
        )
        reply_index += 1


def reddit_thread_to_markdown(
    payload: object,
    include_authors: bool,
    max_comments: int,
) -> tuple[str, str, dict[str, str]]:
    if not isinstance(payload, list) or len(payload) < 2:
        raise ValueError("Reddit returned an unexpected JSON shape.")

    post_children = payload[0].get("data", {}).get("children", [])
    if not post_children:
        raise ValueError("Reddit JSON did not include a post.")

    post_data = post_children[0].get("data", {})
    title = reddit_text(post_data.get("title")) or "Reddit Thread"
    subreddit = reddit_text(post_data.get("subreddit"))
    post_id = reddit_text(post_data.get("id"))
    post_body = reddit_body_text(post_data.get("selftext"))
    post_url = reddit_permalink(post_data.get("permalink"))

    blocks: list[str] = []
    overview = ["## Thread Metadata"]
    if subreddit:
        overview.append(f"- Subreddit: r/{subreddit}")
    if post_id:
        overview.append(f"- Reddit post id: {post_id}")
    score = post_data.get("score")
    if isinstance(score, int):
        overview.append(f"- Post score: {score}")
    comment_count = post_data.get("num_comments")
    if isinstance(comment_count, int):
        overview.append(f"- Reported comment count: {comment_count}")
    if post_url:
        overview.append(f"- Thread permalink: {post_url}")
    blocks.append("\n\n".join(overview))

    post_section = ["## Original Post"]
    post_section.extend(reddit_metadata_lines(post_data, include_authors))
    post_section.append(post_body or "No original post body was available in the Reddit JSON.")
    blocks.append("\n\n".join(post_section))

    comment_children = payload[1].get("data", {}).get("children", [])
    remaining = [max_comments]
    top_level_index = 1
    for child in comment_children:
        if child.get("kind") != "t1":
            continue

        append_reddit_comment(
            child,
            blocks,
            str(top_level_index),
            0,
            remaining,
            include_authors,
        )
        top_level_index += 1

        if remaining[0] <= 0:
            break

    collected_count = max_comments - remaining[0]
    if collected_count == 0:
        blocks.append("## Comments\n\nNo readable comments were available in the Reddit JSON.")

    metadata = {
        "subreddit": f"r/{subreddit}" if subreddit else "",
        "reddit_post_id": post_id,
        "comments_collected": str(collected_count),
    }
    return title, "\n\n".join(blocks), metadata


def markdown_from_reddit_html_fragment(fragment: Tag | None) -> str:
    if not fragment:
        return ""

    blocks: list[str] = []
    for node in fragment.find_all(["p", "li", "blockquote", "pre"], recursive=True):
        text = node.get_text("\n" if node.name == "pre" else " ", strip=True)
        text = reddit_body_text(text)
        if not text:
            continue

        if node.name == "li":
            blocks.append(f"- {text}")
        elif node.name == "blockquote":
            blocks.append(f"> {text}")
        else:
            blocks.append(text)

    if not blocks:
        return reddit_body_text(fragment.get_text("\n", strip=True))

    return "\n\n".join(deduplicate_blocks(blocks))


def old_reddit_score(thing: Tag) -> str:
    score = thing.select_one(".score.unvoted, .score.likes, .score.dislikes")
    if score:
        return clean_inline_text(score.get_text(" ", strip=True))

    return ""


def old_reddit_permalink(thing: Tag) -> str:
    for link in thing.select("a"):
        if clean_inline_text(link.get_text(" ", strip=True)).casefold() == "permalink":
            href = link.get("href", "")
            if isinstance(href, str):
                return reddit_permalink(href)

    permalink = thing.get("data-permalink")
    if isinstance(permalink, str):
        return reddit_permalink(permalink)

    return ""


def old_reddit_metadata_lines(thing: Tag, include_authors: bool) -> list[str]:
    lines: list[str] = []

    if include_authors:
        author = thing.select_one(".author")
        if author:
            author_text = reddit_text(author.get_text(" ", strip=True))
            if author_text:
                lines.append(f"- Author: u/{author_text}")

    score = old_reddit_score(thing)
    if score:
        lines.append(f"- Score: {score}")

    permalink = old_reddit_permalink(thing)
    if permalink:
        lines.append(f"- Permalink: {permalink}")

    return lines


def reddit_thread_html_to_markdown(
    html: str,
    source_url: str,
    include_authors: bool,
    max_comments: int,
) -> tuple[str, str, dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    title_tag = soup.select_one("a.title, .title a, title")
    title = (
        clean_inline_text(title_tag.get_text(" ", strip=True))
        if title_tag
        else title_from_url(source_url)
    )
    title = re.sub(r"\s*:\s*reddit\.com\s*$", "", title, flags=re.IGNORECASE)

    subreddit = ""
    subreddit_link = soup.select_one("a.subreddit")
    if subreddit_link:
        subreddit = reddit_text(subreddit_link.get_text(" ", strip=True))

    blocks: list[str] = []
    overview = ["## Thread Metadata"]
    if subreddit:
        overview.append(f"- Subreddit: {subreddit}")
    overview.append(f"- Thread permalink: {source_url}")
    overview.append("- Capture note: Reddit JSON was unavailable, so this was parsed from old.reddit.com HTML.")
    blocks.append("\n\n".join(overview))

    post = soup.select_one(".thing.link")
    post_section = ["## Original Post"]
    if post:
        post_section.extend(old_reddit_metadata_lines(post, include_authors))
        post_text = markdown_from_reddit_html_fragment(post.select_one(".usertext-body .md"))
        post_section.append(post_text or "No original post body was available in the Reddit HTML.")
    else:
        post_section.append("No original post body was available in the Reddit HTML.")
    blocks.append("\n\n".join(post_section))

    collected_count = 0
    comments = soup.select(".thing.comment")
    for comment in comments:
        if collected_count >= max_comments:
            break

        body = markdown_from_reddit_html_fragment(comment.select_one(".usertext-body .md"))
        if not body:
            continue

        collected_count += 1
        depth_classes = comment.get("class", [])
        depth = 0
        for class_name in depth_classes:
            if isinstance(class_name, str) and class_name.startswith("depth-"):
                try:
                    depth = int(class_name.removeprefix("depth-")) - 1
                except ValueError:
                    depth = 0
                break

        heading_level = min(3 + max(depth, 0), 6)
        section = [f"{'#' * heading_level} Comment {collected_count}"]
        section.extend(old_reddit_metadata_lines(comment, include_authors))
        section.append(body)
        blocks.append("\n\n".join(section))

    if collected_count == 0:
        blocks.append("## Comments\n\nNo readable comments were available in the Reddit HTML.")

    metadata = {
        "subreddit": subreddit,
        "comments_collected": str(collected_count),
        "reddit_capture_fallback": "old_reddit_html",
    }
    return title, "\n\n".join(blocks), metadata


def html_table_to_rows(table: Tag) -> tuple[list[str], list[list[str]]]:
    parsed_rows: list[list[str]] = []

    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        values = [clean_inline_text(cell.get_text(" ", strip=True)) for cell in cells]
        if any(values):
            parsed_rows.append(values)

    if len(parsed_rows) < 2:
        return [], []

    headers = normalize_headers(parsed_rows[0])
    rows = [normalize_row(row, len(headers)) for row in parsed_rows[1:]]
    return headers, rows


def markdown_from_tables(soup: BeautifulSoup) -> str:
    table_blocks: list[str] = []

    for index, table in enumerate(soup.find_all("table"), start=1):
        headers, rows = html_table_to_rows(table)
        if not headers or not rows:
            continue

        caption = table.find("caption")
        title = clean_inline_text(caption.get_text(" ", strip=True)) if caption else f"Table {index}"
        table_markdown = rows_to_markdown(headers, rows, title)
        if table_markdown:
            table_blocks.append(table_markdown)

    return "\n\n".join(table_blocks)


def markdown_from_element(element: Tag) -> str:
    blocks: list[str] = []

    for node in element.descendants:
        if not isinstance(node, Tag):
            continue

        name = node.name.lower()
        text = clean_inline_text(node.get_text(" ", strip=True))
        if not text:
            continue

        if name in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            level = min(int(name[1]), 6)
            blocks.append(f"{'#' * level} {text}")
        elif name == "li":
            blocks.append(f"- {text}")
        elif name == "blockquote":
            blocks.append(f"> {text}")
        elif name in {"p", "pre"}:
            blocks.append(text)

    if not blocks:
        text = element.get_text("\n", strip=True)
        blocks = [clean_inline_text(line) for line in text.splitlines()]

    return "\n\n".join(deduplicate_blocks(blocks))


def visible_strings(element: Tag) -> list[str]:
    strings: list[str] = []
    for value in element.stripped_strings:
        text = clean_inline_text(str(value))
        if text:
            strings.append(text)

    return strings


def extract_visible_badges(soup: BeautifulSoup) -> list[str]:
    badges: list[str] = []
    seen: set[str] = set()

    selectors = [
        ".badge",
        "[class*='badge']",
    ]

    for selector in selectors:
        for tag in soup.select(selector):
            text = clean_inline_text(tag.get_text(" ", strip=True))
            if not text:
                continue

            normalized = text.casefold()
            if normalized in seen:
                continue

            badges.append(text)
            seen.add(normalized)

    return badges


def markdown_from_visible_page(soup: BeautifulSoup) -> str:
    blocks: list[str] = []
    badges = extract_visible_badges(soup)

    if badges:
        blocks.append("## Visible Tags / Badges")
        blocks.extend(f"- {badge}" for badge in badges)

    strings = visible_strings(soup.body if soup.body else soup)
    if strings:
        blocks.append("## All Visible Page Text")
        blocks.extend(f"- {text}" for text in strings)

    return "\n\n".join(blocks)


def deduplicate_blocks(blocks: Iterable[str]) -> list[str]:
    cleaned: list[str] = []
    seen_recent: set[str] = set()

    for block in blocks:
        block = clean_inline_text(block)
        if len(block) < 2:
            continue

        normalized = block.casefold()
        if normalized in seen_recent:
            continue

        cleaned.append(block)
        seen_recent.add(normalized)

        if len(seen_recent) > 250:
            seen_recent = {item.casefold() for item in cleaned[-100:]}

    return cleaned


def slugify(value: str, max_length: int = 80) -> str:
    value = value.lower()
    value = re.sub(r"https?://", "", value)
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = value.strip("-")
    return (value[:max_length].strip("-") or "source")


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    parent = path.parent

    counter = 2
    while True:
        candidate = parent / f"{stem}-{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def build_markdown(
    url: str,
    title: str,
    body_text: str,
    capture_method: str,
    source_type: str = "web_page",
    extra_metadata: dict[str, str] | None = None,
) -> str:
    collected_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    safe_title = title.replace('"', "'")
    safe_url = url.replace('"', "%22")
    safe_source_type = source_type.replace('"', "'")
    front_matter = [
        f'title: "{safe_title}"',
        f'source_url: "{safe_url}"',
        f'source_type: "{safe_source_type}"',
        f'collected_at: "{collected_at}"',
        f'capture_method: "{capture_method}"',
    ]

    for key, value in (extra_metadata or {}).items():
        if not value:
            continue
        safe_key = re.sub(r"[^a-zA-Z0-9_]+", "_", key).strip("_")
        safe_value = value.replace('"', "'")
        front_matter.append(f'{safe_key}: "{safe_value}"')

    return (
        "---\n"
        + "\n".join(front_matter)
        + "\n---\n\n"
        + f"# {title}\n\n"
        + f"{body_text.strip()}\n"
    )


def extract_url_content(
    url: str,
    title_override: str | None,
    reddit_max_comments: int,
    reddit_sort: str,
    include_reddit_authors: bool,
    include_visible_text: bool = False,
) -> ScrapedContent:
    extra_metadata: dict[str, str] = {}

    if is_reddit_url(url):
        fetch_target = reddit_json_url(url, reddit_max_comments, reddit_sort)
        reddit_headers = {
            "User-Agent": REDDIT_USER_AGENT,
            "Accept": "application/json",
        }
        json_error: Exception | None = None
        title = ""
        body_text = ""
        source_type = "reddit_thread"
        capture_method = "scrape_web:reddit_json"

        try:
            response = fetch_url(fetch_target, reddit_headers)
            payload = response.json()
            title, body_text, extra_metadata = reddit_thread_to_markdown(
                payload,
                include_reddit_authors,
                reddit_max_comments,
            )
            capture_method = "scrape_web:reddit_json"
        except (requests.RequestException, ValueError) as public_json_error:
            json_error = public_json_error
            try:
                token = reddit_oauth_token()
            except (requests.RequestException, ValueError) as token_error:
                token = None
                json_error = token_error

            if token:
                oauth_headers = {
                    "User-Agent": REDDIT_USER_AGENT,
                    "Accept": "application/json",
                    "Authorization": f"Bearer {token}",
                }
                try:
                    response = fetch_url(
                        reddit_oauth_json_url(url, reddit_max_comments, reddit_sort),
                        oauth_headers,
                    )
                    payload = response.json()
                    title, body_text, extra_metadata = reddit_thread_to_markdown(
                        payload,
                        include_reddit_authors,
                        reddit_max_comments,
                    )
                    capture_method = "scrape_web:reddit_oauth_json"
                except (requests.RequestException, ValueError) as oauth_error:
                    json_error = oauth_error

            if not body_text:
                fallback_url = reddit_old_html_url(url, reddit_sort)
                html_headers = {
                    "User-Agent": REDDIT_USER_AGENT,
                    "Accept": "text/html,application/xhtml+xml",
                }
                try:
                    response = fetch_url(fallback_url, html_headers)
                except requests.RequestException as html_error:
                    raise ValueError(
                        "Reddit blocked public JSON, OAuth JSON was not configured or "
                        "did not work, and old.reddit.com HTML was also blocked. Add "
                        "REDDIT_CLIENT_ID and REDDIT_CLIENT_SECRET to .env, or save "
                        "the thread manually as Markdown."
                    ) from html_error

                title, body_text, extra_metadata = reddit_thread_html_to_markdown(
                    response.text,
                    url,
                    include_reddit_authors,
                    reddit_max_comments,
                )
                extra_metadata["reddit_json_error"] = str(json_error).replace("\n", " ")
                capture_method = "scrape_web:old_reddit_html"

        title = title_override or title
    else:
        google_csv_url = google_sheets_csv_url(url)
        fetch_target = google_csv_url or url
        response = fetch_url(fetch_target)
        content_type = response.headers.get("content-type", "")
        source_type = "web_page"
        capture_method = "scrape_web"

        if google_csv_url or looks_like_tabular_response(fetch_target, content_type):
            source_type = "tabular_data"
            capture_method = "scrape_web:tabular"
            title = title_override or title_from_url(url)

            if urlparse(fetch_target).path.lower().endswith(".xlsx") or "spreadsheetml" in content_type.lower():
                headers, rows = parse_xlsx_rows(response.content)
            else:
                text = decode_text_response(response)
                delimiter = "\t" if "tab-separated-values" in content_type.lower() else None
                headers, rows = parse_delimited_rows(text, delimiter=delimiter)

            body_text = rows_to_markdown(headers, rows)
        else:
            if "html" not in content_type.lower():
                raise ValueError(
                    f"Expected an HTML or tabular source, but got content type: "
                    f"{content_type or 'unknown'}"
                )

            soup = BeautifulSoup(response.text, "html.parser")
            clean_soup(soup, keep_visible_chrome=include_visible_text)

            title = title_override or extract_title(soup, url)
            if include_visible_text:
                body_text = markdown_from_visible_page(soup)
                capture_method = "scrape_web:visible_text"
            else:
                table_text = markdown_from_tables(soup)
                if table_text:
                    source_type = "html_table"
                    capture_method = "scrape_web:html_table"
                    body_text = table_text
                else:
                    content_root = choose_content_root(soup)
                    body_text = markdown_from_element(content_root)

    if len(body_text) < 200:
        raise ValueError(
            "Extracted less than 200 characters. This page may be JavaScript-rendered, "
            "blocked, empty, or better collected manually."
        )

    return ScrapedContent(
        title=title,
        body_text=body_text,
        source_type=source_type,
        capture_method=capture_method,
        extra_metadata=extra_metadata,
    )


def paginated_url(
    base_url: str,
    page_number: int,
    page_param: str,
    first_page_number: int,
) -> str:
    if page_number == first_page_number:
        return base_url

    parsed = urlparse(base_url)
    query_pairs = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if key != page_param
    ]
    query_pairs.append((page_param, str(page_number)))

    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            parsed.params,
            urlencode(query_pairs),
            parsed.fragment,
        )
    )


def discover_links_by_text(url: str, link_text: str) -> list[str]:
    response = fetch_url(url)
    if "html" not in response.headers.get("content-type", "").lower():
        raise ValueError(f"Expected an HTML listing page for link discovery: {url}")

    soup = BeautifulSoup(response.text, "html.parser")
    expected = clean_inline_text(link_text).casefold()
    links: list[str] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a"):
        text = clean_inline_text(anchor.get_text(" ", strip=True)).casefold()
        href = anchor.get("href")
        if not isinstance(href, str) or not href:
            continue
        if text != expected:
            continue

        absolute_url = urljoin(url, href)
        if absolute_url in seen:
            continue

        links.append(absolute_url)
        seen.add(absolute_url)

    return links


def output_path_for(url: str, title: str, output: Path | None) -> Path:
    DOCUMENTS_DIR.mkdir(exist_ok=True)

    if output is None:
        parsed = urlparse(url)
        source_hint = f"{parsed.netloc}-{title}"
        output = DOCUMENTS_DIR / f"{slugify(source_hint)}.md"
    elif output.suffix.lower() != ".md":
        output = output.with_suffix(".md")

    return unique_path(output)


def write_markdown_file(
    url: str,
    title: str,
    body_text: str,
    capture_method: str,
    source_type: str,
    extra_metadata: dict[str, str],
    output: Path | None,
) -> Path:
    output = output_path_for(url, title, output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        build_markdown(
            url,
            title,
            body_text,
            capture_method,
            source_type,
            extra_metadata,
        ),
        encoding="utf-8",
    )

    return output


def scrape_url_to_markdown(
    url: str,
    output: Path | None,
    title_override: str | None,
    reddit_max_comments: int,
    reddit_sort: str,
    include_reddit_authors: bool,
    include_visible_text: bool,
) -> Path:
    content = extract_url_content(
        url,
        title_override,
        reddit_max_comments,
        reddit_sort,
        include_reddit_authors,
        include_visible_text,
    )

    return write_markdown_file(
        url,
        content.title,
        content.body_text,
        content.capture_method,
        content.source_type,
        content.extra_metadata,
        output,
    )


def scrape_paginated_to_markdown(
    url: str,
    output: Path | None,
    title_override: str | None,
    page_start: int,
    page_end: int,
    page_param: str,
    follow_read_more: bool,
    read_more_text: str,
    reddit_max_comments: int,
    reddit_sort: str,
    include_reddit_authors: bool,
    include_visible_text: bool,
) -> Path:
    if page_start < 1:
        raise ValueError("--page-start must be 1 or greater.")
    if page_end < page_start:
        raise ValueError("--page-end must be greater than or equal to --page-start.")

    page_sections: list[str] = []
    page_urls: list[str] = []
    detail_urls: list[str] = []
    first_title = title_override

    for page_number in range(page_start, page_end + 1):
        current_url = paginated_url(url, page_number, page_param, page_start)
        page_urls.append(current_url)

        if follow_read_more:
            links = discover_links_by_text(current_url, read_more_text)
            page_blocks = [
                f"## Listing Page {page_number}",
                f"- Listing page URL: {current_url}",
                f"- Detail links found: {len(links)}",
            ]

            for detail_index, detail_url in enumerate(links, start=1):
                if detail_url in detail_urls:
                    continue

                detail_urls.append(detail_url)
                content = extract_url_content(
                    detail_url,
                    None,
                    reddit_max_comments,
                    reddit_sort,
                    include_reddit_authors,
                    True,
                )

                if first_title is None:
                    first_title = content.title

                page_blocks.append(
                    "\n\n".join(
                        [
                            f"### Detail {detail_index}: {content.title}",
                            f"- Detail URL: {detail_url}",
                            content.body_text,
                        ]
                    )
                )

            page_sections.append("\n\n".join(page_blocks))
        else:
            content = extract_url_content(
                current_url,
                None,
                reddit_max_comments,
                reddit_sort,
                include_reddit_authors,
                include_visible_text,
            )

            if first_title is None:
                first_title = content.title

            page_sections.append(
                "\n\n".join(
                    [
                        f"## Page {page_number}",
                        f"- Page URL: {current_url}",
                        f"- Page title: {content.title}",
                        content.body_text,
                    ]
                )
            )

    title = title_override or f"{first_title} (pages {page_start}-{page_end})"
    body_text = "\n\n".join(page_sections)
    extra_metadata = {
        "page_start": str(page_start),
        "page_end": str(page_end),
        "page_param": page_param,
        "pages_collected": str(len(page_urls)),
    }
    source_type = "paginated_web_pages"
    capture_method = "scrape_web:pagination"

    if follow_read_more:
        source_type = "paginated_detail_pages"
        capture_method = "scrape_web:pagination_read_more"
        extra_metadata["followed_link_text"] = read_more_text
        extra_metadata["detail_pages_collected"] = str(len(detail_urls))

    return write_markdown_file(
        url,
        title,
        body_text,
        capture_method,
        source_type,
        extra_metadata,
        output,
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="scrape-web",
        description=(
            "Scrape web, Reddit, table, or paginated URLs and save cleaned "
            "Markdown source documents."
        )
    )
    parser.add_argument("url", help="The web URL to scrape.")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        help="Optional output path. Defaults to documents/<page-title>.md.",
    )
    parser.add_argument(
        "--title",
        help="Optional title override for the saved Markdown metadata and heading.",
    )
    parser.add_argument(
        "--page-start",
        type=int,
        default=1,
        help="First page number to scrape when using paginated mode. Default: 1.",
    )
    parser.add_argument(
        "--page-end",
        type=int,
        help=(
            "Last page number to scrape. When provided, the script scrapes "
            "all pages from --page-start through --page-end into one Markdown file."
        ),
    )
    parser.add_argument(
        "--page-param",
        default="page",
        help="Query parameter used for pagination. Default: page.",
    )
    parser.add_argument(
        "--follow-read-more",
        action="store_true",
        help=(
            "For paginated listing/card pages, follow each page's Read more "
            "links and save the linked detail pages instead of listing teasers."
        ),
    )
    parser.add_argument(
        "--read-more-text",
        default="Read more",
        help="Visible link text to follow when using --follow-read-more. Default: Read more.",
    )
    parser.add_argument(
        "--include-visible-text",
        action="store_true",
        help=(
            "Preserve all visible text nodes from HTML pages, including tags, "
            "badges, statuses, navigation labels, and action text."
        ),
    )
    parser.add_argument(
        "--reddit-max-comments",
        type=int,
        default=30,
        help="Maximum readable Reddit comments/replies to store. Default: 30.",
    )
    parser.add_argument(
        "--reddit-sort",
        choices=["confidence", "top", "new", "controversial", "old", "qa"],
        default="top",
        help="Comment sort order for Reddit thread JSON. Default: top.",
    )
    parser.add_argument(
        "--include-reddit-authors",
        action="store_true",
        help="Include Reddit usernames in saved Markdown. Default: omit usernames.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    load_dotenv()
    args = parse_args(sys.argv[1:] if argv is None else argv)

    try:
        if args.page_end is not None:
            saved_path = scrape_paginated_to_markdown(
                args.url,
                args.output,
                args.title,
                args.page_start,
                args.page_end,
                args.page_param,
                args.follow_read_more,
                args.read_more_text,
                args.reddit_max_comments,
                args.reddit_sort,
                args.include_reddit_authors,
                args.include_visible_text,
            )
        else:
            saved_path = scrape_url_to_markdown(
                args.url,
                args.output,
                args.title,
                args.reddit_max_comments,
                args.reddit_sort,
                args.include_reddit_authors,
                args.include_visible_text,
            )
    except Exception as exc:
        print(f"Failed to scrape URL: {exc}", file=sys.stderr)
        return 1

    print(f"Saved scraped document to: {saved_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
